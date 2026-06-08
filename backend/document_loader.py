from __future__ import annotations

import csv
import html
import io
import json
from html.parser import HTMLParser
from pathlib import Path


EXCEL_SUFFIXES = {".xlsx", ".xlsm", ".xltx", ".xltm", ".xls"}
HTML_SUFFIXES = {".html", ".htm"}
TEXT_SUFFIXES = {
    ".txt",
    ".md",
    ".markdown",
    ".py",
    ".js",
    ".ts",
    ".tsx",
    ".jsx",
    ".css",
    ".xml",
    ".yaml",
    ".yml",
    ".log",
}


def extract_text(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        return _read_pdf(path)
    if suffix == ".docx":
        return _read_docx(path)
    if suffix == ".csv":
        return _read_csv(path)
    if suffix in EXCEL_SUFFIXES:
        return _read_excel(path)
    if suffix in HTML_SUFFIXES:
        return _read_html(path)
    if suffix == ".json":
        return _read_json(path)
    if suffix in TEXT_SUFFIXES:
        return _read_text(path)
    return _read_text(path)


def chunk_text(text: str, chunk_size: int = 1200, overlap: int = 180) -> list[str]:
    cleaned = "\n".join(line.strip() for line in text.splitlines())
    cleaned = "\n".join(line for line in cleaned.splitlines() if line)
    if not cleaned:
        return []

    chunks: list[str] = []
    start = 0
    text_length = len(cleaned)
    while start < text_length:
        end = min(start + chunk_size, text_length)
        window = cleaned[start:end]
        split_at = max(window.rfind("\n\n"), window.rfind(". "), window.rfind("? "), window.rfind("! "))
        if split_at > chunk_size * 0.55 and end < text_length:
            end = start + split_at + 1
        chunk = cleaned[start:end].strip()
        if chunk:
            chunks.append(chunk)
        if end >= text_length:
            break
        start = max(0, end - overlap)
    return chunks


def _read_text(path: Path) -> str:
    data = path.read_bytes()
    for encoding in ("utf-8", "utf-16", "cp1252", "latin-1"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="ignore")


def _read_pdf(path: Path) -> str:
    try:
        import fitz

        with fitz.open(str(path)) as document:
            return "\n\n".join(page.get_text() for page in document)
    except ImportError:
        pass

    try:
        from pypdf import PdfReader
    except ImportError as exc:
        raise RuntimeError("PDF support needs PyMuPDF or pypdf. Install backend/requirements.txt.") from exc

    reader = PdfReader(str(path))
    pages = [page.extract_text() or "" for page in reader.pages]
    return "\n\n".join(pages)


def _read_docx(path: Path) -> str:
    try:
        from docx import Document
    except ImportError as exc:
        raise RuntimeError("DOCX support needs python-docx. Install backend/requirements.txt.") from exc

    document = Document(str(path))
    paragraphs = [paragraph.text for paragraph in document.paragraphs if paragraph.text.strip()]
    return "\n".join(paragraphs)


def _read_csv(path: Path) -> str:
    text = _read_text(path)
    rows = csv.reader(io.StringIO(text))
    return "\n".join(" | ".join(cell.strip() for cell in row) for row in rows)


def _read_excel(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix == ".xls":
        return _read_xls(path)
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Excel support needs openpyxl. Install backend/requirements.txt.") from exc

    workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    sheet_text: list[str] = []
    for sheet in workbook.worksheets:
        rows = []
        for row in sheet.iter_rows(values_only=True):
            values = ["" if value is None else str(value).strip() for value in row]
            if any(values):
                rows.append(values)
        if rows:
            sheet_text.append(f"SHEET: {sheet.title}\n{_render_delimited_rows(rows)}")
    workbook.close()
    return "\n\n".join(sheet_text)


def _read_xls(path: Path) -> str:
    try:
        import xlrd
    except ImportError as exc:
        raise RuntimeError("Legacy .xls support needs xlrd. Install backend/requirements.txt.") from exc

    workbook = xlrd.open_workbook(str(path))
    sheet_text: list[str] = []
    for sheet in workbook.sheets():
        rows = []
        for row_index in range(sheet.nrows):
            values = [str(sheet.cell_value(row_index, col_index)).strip() for col_index in range(sheet.ncols)]
            if any(values):
                rows.append(values)
        if rows:
            sheet_text.append(f"SHEET: {sheet.name}\n{_render_delimited_rows(rows)}")
    return "\n\n".join(sheet_text)


def _read_html(path: Path) -> str:
    parser = _ReadableHTMLParser()
    parser.feed(_read_text(path))
    return parser.text()


def _read_json(path: Path) -> str:
    payload = json.loads(_read_text(path))
    return json.dumps(payload, ensure_ascii=False, indent=2)


def _render_delimited_rows(rows: list[list[str]]) -> str:
    return "\n".join(" | ".join(str(cell).strip() for cell in row) for row in rows)


class _ReadableHTMLParser(HTMLParser):
    BLOCK_TAGS = {
        "address",
        "article",
        "aside",
        "blockquote",
        "br",
        "caption",
        "div",
        "figcaption",
        "figure",
        "footer",
        "h1",
        "h2",
        "h3",
        "h4",
        "h5",
        "h6",
        "header",
        "li",
        "main",
        "p",
        "pre",
        "section",
        "table",
        "td",
        "th",
        "tr",
    }
    SKIP_TAGS = {"script", "style", "noscript", "svg"}

    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.parts: list[str] = []
        self.skip_depth = 0

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag in self.SKIP_TAGS:
            self.skip_depth += 1
            return
        if tag in self.BLOCK_TAGS:
            self.parts.append("\n")

    def handle_endtag(self, tag: str) -> None:
        if tag in self.SKIP_TAGS and self.skip_depth:
            self.skip_depth -= 1
            return
        if tag in self.BLOCK_TAGS:
            self.parts.append("\n")

    def handle_data(self, data: str) -> None:
        if self.skip_depth:
            return
        text = html.unescape(data).strip()
        if text:
            self.parts.append(text)
            self.parts.append(" ")

    def text(self) -> str:
        lines = []
        for line in "".join(self.parts).splitlines():
            cleaned = " ".join(line.split())
            if cleaned:
                lines.append(cleaned)
        return "\n".join(lines)
